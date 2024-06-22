# %% [markdown]
"""

# 練習問題解答


## 第1章 練習問題

### 1.1
次の統計について、上2桁の概数で答えよ（いずれも、本年度、ないしは直近の数字）。

(i) 日本の人口

"""

# %%
"""
e-StatをAPI利用する際は、登録しないとダメなのね。
API利用は改めて行うとして、とりあえずHPからデータをダウンロードするか。
https://www.e-stat.go.jp/stat-search/files?page=1&layout=datalist&toukei=00200524&tstat=000000090001&cycle=7&year=20220&month=0&tclass1=000001011679&tclass2val=0
"""


import openpyxl as op

# Excel ファイル（ワークブック）の読み込み
wb = op.load_workbook("./a00100.xlsx")

# ワークシートの有効化
ws = wb.active
value = ws["K10"].value


population = int(value)*1000
print('Population of JAPAN is',f"{population:,}",'人(',ws["I10"].value,')')


# %% [markdown]
"""

(ii) 日本の国家予算（一般会計）
1. 日本の国民総生産


"""
# %%

print('Hello World!')



